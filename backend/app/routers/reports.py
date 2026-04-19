"""
Reports router - /api/statistics, /api/reports, etc.
"""
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from app.services.report_service import ReportService
from app.services.material_service import MaterialService
from app.deps import require_admin
import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json

router = APIRouter(tags=["reports"])
report_service = ReportService()
material_service = MaterialService()


@router.get("/statistics")
async def get_statistics(admin: dict = Depends(require_admin)):
    """Get dashboard statistics."""
    return await report_service.get_statistics()


@router.get("/reports")
async def get_reports(admin: dict = Depends(require_admin)):
    """Get detailed reports."""
    return await report_service.get_reports()


@router.get("/alerts/low-stock")
async def get_low_stock(admin: dict = Depends(require_admin)):
    """Get low stock alerts."""
    return await material_service.get_low_stock()


@router.get("/reports/export-orders")
async def export_orders(admin: dict = Depends(require_admin)):
    """Export orders to Excel."""
    orders = await report_service.get_orders_for_export()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar"
    
    # Styling
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='6C63FF', end_color='6C63FF', fill_type='solid')
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    
    headers = ['#', 'Buyurtma kodi', 'Diler', 'Mahsulotlar', 'Jami kv.m', 'Jami narx ($)', 'Status', 'Sana']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    status_map = {
        "kutilmoqda": "Kutilmoqda",
        "tasdiqlangan": "Tasdiqlangan",
        "tayyorlanmoqda": "Tayyorlanmoqda",
        "tayyor": "Tayyor",
        "yetkazilmoqda": "Yetkazilmoqda",
        "yetkazildi": "Yetkazildi",
        "rad_etilgan": "Rad etilgan"
    }
    
    for idx, order in enumerate(orders, 1):
        items = json.loads(order["items"]) if isinstance(order["items"], str) else order["items"]
        item_names = ", ".join([f'{it["material_name"]} ({it.get("width", 0)}x{it.get("height", 0)}m)' for it in items])
        
        row = [
            idx,
            order["order_code"],
            order["dealer_name"],
            item_names,
            round(order["total_sqm"], 2),
            round(order["total_price"], 2),
            status_map.get(order["status"], order["status"]),
            order["created_at"][:16].replace("T", " ")
        ]
        
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=idx+1, column=col, value=val)
            cell.border = border
            if col in [5, 6]:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    today = datetime.now(timezone.utc).strftime('%Y%m%d')
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=buyurtmalar_{today}.xlsx"}
    )
